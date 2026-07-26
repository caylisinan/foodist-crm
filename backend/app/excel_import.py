"""
Excel (XLSX) içe aktarma. Kullanıcı önce sütun başlıklarını görür,
sistem alanlarıyla eşleştirir, sonra kayıtlar veritabanına yazılır.

Not: Eski .xls formatı için openpyxl yeterli değildir (xlrd gerekir).
MVP kapsamında .xlsx odaklanılmıştır; .xls dosyaları Excel'de
"Farklı Kaydet → xlsx" ile kolayca dönüştürülebilir.
"""
import os
import uuid
import tempfile
from openpyxl import load_workbook

TEMP_DIR = os.path.join(tempfile.gettempdir(), "foodist_crm_imports")
os.makedirs(TEMP_DIR, exist_ok=True)

# Sistem tarafındaki hedef alanlar (entity_type'a göre)
TARGET_FIELDS = {
    "buyer": [
        "company_name", "contact_name", "country", "contact_email",
        "contact_phone", "sector", "interested_products",
        "max_meetings", "max_minutes", "notes",
    ],
    "participant": [
        "company_name", "contact_name", "country", "contact_email",
        "contact_phone", "sector", "offered_products", "stand_no", "notes",
    ],
}

TARGET_FIELD_LABELS = {
    "company_name": "Firma Adı",
    "contact_name": "Yetkili Ad Soyad",
    "country": "Ülke",
    "contact_email": "E-posta",
    "contact_phone": "Telefon",
    "sector": "Sektör",
    "interested_products": "İlgilenilen Ürünler",
    "offered_products": "Sunulan Ürünler",
    "stand_no": "Stand No",
    "max_meetings": "Maks. Toplantı",
    "max_minutes": "Maks. Dakika",
    "notes": "Not",
}


def save_uploaded_file(file_bytes: bytes, original_filename: str) -> str:
    """Yüklenen dosyayı geçici bir konuma kaydeder, bir token döner."""
    token = str(uuid.uuid4())
    ext = os.path.splitext(original_filename)[1] or ".xlsx"
    path = os.path.join(TEMP_DIR, token + ext)
    with open(path, "wb") as f:
        f.write(file_bytes)
    return token, path


def _find_file_by_token(token: str) -> str:
    for fname in os.listdir(TEMP_DIR):
        if fname.startswith(token):
            return os.path.join(TEMP_DIR, fname)
    raise FileNotFoundError("Yüklenen dosya bulunamadı, lütfen tekrar yükleyin.")


def preview_excel(token: str, max_rows: int = 5) -> dict:
    path = _find_file_by_token(token)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, [])
    header = [str(h) if h is not None else "" for h in header]
    preview_rows = []
    for i, row in enumerate(rows_iter):
        if i >= max_rows:
            break
        preview_rows.append([("" if v is None else v) for v in row])
    return {"columns": header, "preview_rows": preview_rows}


def apply_mapping(token: str, mapping: dict) -> list[dict]:
    """
    mapping: {"company_name": "Buyer Name", "country": "Country", ...}
    Excel'deki kolon adına göre veriyi okuyup sistem alanı isimli
    dict listesi döner.
    """
    path = _find_file_by_token(token)
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, [])
    header = [str(h) if h is not None else "" for h in header]
    col_index = {name: idx for idx, name in enumerate(header)}

    records = []
    for row in rows_iter:
        if row is None or all(v is None for v in row):
            continue
        record = {}
        for target_field, excel_col in mapping.items():
            if not excel_col or excel_col not in col_index:
                record[target_field] = None
                continue
            idx = col_index[excel_col]
            value = row[idx] if idx < len(row) else None
            record[target_field] = value
        # en azından firma adı olmalı
        if record.get("company_name"):
            records.append(record)
    return records
